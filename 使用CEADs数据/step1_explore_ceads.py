"""
探索CEADs数据文件结构
"""
import pandas as pd
import numpy as np
from openpyxl import load_workbook

# 先用openpyxl查看工作表结构
file_path = '1997-2019年290个中国城市碳排放清单 (1).xlsx'
wb = load_workbook(file_path)

print("=" * 80)
print("CEADs Excel工作表探索")
print("=" * 80)

print(f"\n工作表列表: {wb.sheetnames}")

# 尝试读取每个工作表的前几行
for sheet_name in wb.sheetnames:
    print(f"\n{'='*80}")
    print(f"工作表: {sheet_name}")
    print(f"{'='*80}")

    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        print(f"形状: {df.shape}")
        print(f"列名: {df.columns.tolist()[:10]}")  # 只显示前10列
        print(f"\n前5行:")
        print(df.head())
    except Exception as e:
        print(f"读取失败: {e}")
